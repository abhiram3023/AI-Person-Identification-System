import tkinter as tk
from tkinter import ttk
from tkinter import messagebox as mess
import tkinter.simpledialog as tsd
import cv2
import os
import numpy as np
from PIL import Image, ImageTk
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime
import time
import shutil
import platform
from subprocess import check_output, CalledProcessError


# ========================= PATHS =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TRAINING_IMAGE_DIR = os.path.join(BASE_DIR, "TrainingImage")
TRAINING_LABEL_DIR = os.path.join(BASE_DIR, "TrainingImageLabel")
PERSON_DETAILS_DIR = os.path.join(BASE_DIR, "PersonDetails")
IDENTIFICATION_DIR = os.path.join(BASE_DIR, "IdentificationRecords")

PERSON_DETAILS_FILE = os.path.join(PERSON_DETAILS_DIR, "PersonDetails.xlsx")
IDENTIFICATION_FILE = os.path.join(IDENTIFICATION_DIR, "IdentificationLog.xlsx")
PASSWORD_FILE = os.path.join(TRAINING_LABEL_DIR, "psd.txt")
TRAINER_FILE = os.path.join(TRAINING_LABEL_DIR, "Trainer.yml")
CASCADE_FILE = os.path.join(BASE_DIR, "haarcascade_frontalface_default.xml")


# ========================= THEME =========================
BG = "#eaf1f8"
HEADER_BG = "#16324f"
HEADER_FG = "#ffffff"
CARD_BG = "#ffffff"
CARD_BORDER = "#cbd5e1"
ACCENT = "#2563eb"
ACCENT_DARK = "#1d4ed8"
SUCCESS = "#16a34a"
WARNING = "#dc2626"
MUTED = "#475569"
TABLE_BG = "#f8fafc"
TABLE_HEADER = "#e2e8f0"


# ========================= HELPERS =========================
def assure_path_exists(path):
    os.makedirs(path, exist_ok=True)


def autofit_excel(file_path):
    try:
        wb = load_workbook(file_path)
    except Exception:
        return

    for ws in wb.worksheets:
        for col in ws.columns:
            try:
                col_letter = get_column_letter(col[0].column)
            except Exception:
                continue

            max_length = 0
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = max(10, max_length + 3)

    try:
        wb.save(file_path)
    except Exception:
        pass


def check_haarcascadefile():
    if os.path.isfile(CASCADE_FILE):
        return True
    mess.showerror(
        "Missing File",
        'Please make sure "haarcascade_frontalface_default.xml" exists in the application folder.',
    )
    return False


def is_capslock_on():
    system = platform.system()
    try:
        if system == "Windows":
            import ctypes
            hll = ctypes.WinDLL("User32.dll")
            VK_CAPITAL = 0x14
            state = hll.GetKeyState(VK_CAPITAL)
            return bool(state & 0x0001)
        elif system == "Darwin":
            try:
                from Quartz import CGEventSourceKeyState, kCGEventSourceStateHIDSystemState
                return bool(CGEventSourceKeyState(kCGEventSourceStateHIDSystemState, 57))
            except Exception:
                return False
        elif system == "Linux":
            try:
                out = check_output(["xset", "q"]).decode(errors="ignore")
                return "Caps Lock:   on" in out or "Caps Lock: on" in out
            except (CalledProcessError, FileNotFoundError):
                return False
        return False
    except Exception:
        return False


def start_capslock_monitor(entry_widget, warning_label, interval=300):
    def poll():
        try:
            if entry_widget.winfo_exists() and warning_label.winfo_exists():
                warning_label.config(text="⚠ Caps Lock is ON!" if is_capslock_on() else "")
                entry_widget.after(interval, poll)
        except Exception:
            pass

    poll()


def ensure_person_details_file():
    assure_path_exists(PERSON_DETAILS_DIR)
    if not os.path.isfile(PERSON_DETAILS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "People"
        ws.append(["SERIAL NO.", "ID", "NAME"])
        wb.save(PERSON_DETAILS_FILE)
        autofit_excel(PERSON_DETAILS_FILE)


def ensure_identification_file():
    assure_path_exists(IDENTIFICATION_DIR)
    if not os.path.isfile(IDENTIFICATION_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Identification"
        ws.append(["S.No", "ID", "Name", "Date", "Time"])
        wb.save(IDENTIFICATION_FILE)
        autofit_excel(IDENTIFICATION_FILE)


def compute_registration_count():
    try:
        if not os.path.isfile(PERSON_DETAILS_FILE):
            return 0
        df = pd.read_excel(PERSON_DETAILS_FILE)
        df = df.dropna(how="all")
        if "ID" in df.columns:
            ids = pd.to_numeric(df["ID"], errors="coerce")
            return int(ids.notna().sum())
        return max(0, len(df) - 1)
    except Exception:
        return 0


def update_registration_counter():
    count = compute_registration_count()
    stats_label.config(text=f"Total Registered People: {count}")


def get_next_serial_from_sheet(ws):
    for r in range(ws.max_row, 0, -1):
        val = ws.cell(row=r, column=1).value
        try:
            return int(val) + 1
        except Exception:
            continue
    return 1


def refresh_identification_tree():
    for item in tv.get_children():
        tv.delete(item)

    if not os.path.isfile(IDENTIFICATION_FILE):
        return

    try:
        df = pd.read_excel(IDENTIFICATION_FILE)
        df = df.dropna(how="all")
        if df.empty:
            return

        for _, row in df.iterrows():
            vals = list(row.values)
            if len(vals) >= 5:
                tv.insert("", "end", text=str(vals[0]), values=(vals[1], vals[2], vals[3], vals[4]))

        # auto-scroll to latest
        children = tv.get_children()
        if children:
            tv.see(children[-1])
    except Exception:
        pass


def append_identification_log(student_id, student_name, date_str, time_str):
    ensure_identification_file()

    try:
        wb = load_workbook(IDENTIFICATION_FILE)
        ws = wb.active
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append(["S.No", "ID", "Name", "Date", "Time"])

    if ws.max_row == 0:
        ws.append(["S.No", "ID", "Name", "Date", "Time"])

    serial_to_write = get_next_serial_from_sheet(ws)
    ws.append([serial_to_write, student_id, student_name, date_str, time_str])
    wb.save(IDENTIFICATION_FILE)
    autofit_excel(IDENTIFICATION_FILE)

    tv.insert("", "end", text=str(serial_to_write), values=(student_id, student_name, date_str, time_str))
    children = tv.get_children()
    if children:
        tv.see(children[-1])


def get_student_row_by_id(df, student_id):
    try:
        ids = pd.to_numeric(df["ID"], errors="coerce")
        matched = df.loc[ids == int(student_id)]
        if not matched.empty:
            return matched.iloc[0]
    except Exception:
        pass
    return None


def capture_face_roi(gray, x, y, w, h):
    roi = gray[y:y + h, x:x + w]
    roi = cv2.resize(roi, (200, 200))
    return roi


# ========================= PASSWORD DIALOG =========================
def ask_password_dialog(parent, title="Password", prompt="Enter Password", require_confirm=False):
    result = {"password": None}
    dialog = tk.Toplevel(parent)
    dialog.transient(parent)
    dialog.grab_set()
    dialog.title(title)
    dialog.geometry("420x220" if require_confirm else "420x150")
    dialog.resizable(False, False)
    dialog.configure(bg=CARD_BG)

    lbl = tk.Label(dialog, text=prompt, bg=CARD_BG, fg=MUTED, font=("Segoe UI", 11, "bold"))
    lbl.place(x=10, y=10)

    pwd_var = tk.StringVar()
    pwd_entry = tk.Entry(dialog, textvariable=pwd_var, width=30, show="*", font=("Segoe UI", 12, "bold"))
    pwd_entry.place(x=10, y=40)
    pwd_entry.focus_set()

    pwd_warning = tk.Label(dialog, text="", bg=CARD_BG, fg=WARNING, font=("Segoe UI", 9, "bold"))
    pwd_warning.place(x=10, y=70)
    start_capslock_monitor(pwd_entry, pwd_warning)

    def toggle_pwd():
        if pwd_entry.cget("show") == "":
            pwd_entry.config(show="*")
            toggle_btn.config(text="Show")
        else:
            pwd_entry.config(show="")
            toggle_btn.config(text="Hide")

    toggle_btn = tk.Button(dialog, text="Show", command=toggle_pwd, width=6)
    toggle_btn.place(x=320, y=38)

    confirm_var = tk.StringVar()
    if require_confirm:
        confirm_lbl = tk.Label(dialog, text="Confirm Password", bg=CARD_BG, fg=MUTED, font=("Segoe UI", 11, "bold"))
        confirm_lbl.place(x=10, y=95)
        confirm_entry = tk.Entry(dialog, textvariable=confirm_var, width=30, show="*", font=("Segoe UI", 12, "bold"))
        confirm_entry.place(x=10, y=125)

        confirm_warning = tk.Label(dialog, text="", bg=CARD_BG, fg=WARNING, font=("Segoe UI", 9, "bold"))
        confirm_warning.place(x=10, y=155)
        start_capslock_monitor(confirm_entry, confirm_warning)

        def toggle_confirm():
            if confirm_entry.cget("show") == "":
                confirm_entry.config(show="*")
                toggle_confirm_btn.config(text="Show")
            else:
                confirm_entry.config(show="")
                toggle_confirm_btn.config(text="Hide")

        toggle_confirm_btn = tk.Button(dialog, text="Show", command=toggle_confirm, width=6)
        toggle_confirm_btn.place(x=320, y=123)

    def on_ok():
        pwd = pwd_var.get()
        if require_confirm:
            conf = confirm_var.get()
            if not pwd:
                mess.showerror("No Password", "Please enter a password.")
                return
            if pwd != conf:
                mess.showerror("Mismatch", "Passwords do not match.")
                return
        result["password"] = pwd
        dialog.destroy()

    def on_cancel():
        dialog.destroy()

    ok_y = 170 if require_confirm else 100
    ok_btn = tk.Button(dialog, text="OK", width=10, command=on_ok)
    ok_btn.place(x=240, y=ok_y)
    cancel_btn = tk.Button(dialog, text="Cancel", width=10, command=on_cancel)
    cancel_btn.place(x=340, y=ok_y)

    dialog.wait_window()
    return result["password"]


# ========================= PASSWORD FLOW =========================
def save_pass():
    assure_path_exists(TRAINING_LABEL_DIR)

    if not os.path.isfile(PASSWORD_FILE):
        new_pas = tsd.askstring("Old Password not found", "Please enter a new password below", show="*")
        if new_pas is None:
            mess.showerror("No Password Entered", "Password not set!! Please try again")
            return
        with open(PASSWORD_FILE, "w", encoding="utf-8") as tf:
            tf.write(new_pas)
        mess.showinfo("Password Registered", "New password was registered successfully!!")
        return

    op = old.get()
    newp = new.get()
    nnewp = nnew.get()

    with open(PASSWORD_FILE, "r", encoding="utf-8") as tf:
        key = tf.read()

    if op != key:
        mess.showerror("Wrong Password", "Please enter correct old password.")
        return

    if newp != nnewp:
        mess.showerror("Error", "Confirm new password again!!!")
        return

    with open(PASSWORD_FILE, "w", encoding="utf-8") as txf:
        txf.write(newp)

    mess.showinfo("Password Changed", "Password changed successfully!!")
    master.destroy()


def change_pass():
    global master, old, new, nnew

    master = tk.Toplevel(window)
    master.geometry("420x200")
    master.resizable(False, False)
    master.title("Change Password")
    master.configure(background=CARD_BG)
    master.grab_set()

    lbl4 = tk.Label(master, text="Enter Old Password", bg=CARD_BG, fg=MUTED, font=("Segoe UI", 12, "bold"))
    lbl4.place(x=10, y=10)
    old = tk.Entry(master, width=22, fg="black", relief="solid", font=("Segoe UI", 12, "bold"), show="*")
    old.place(x=200, y=10)
    old_warning = tk.Label(master, text="", bg=CARD_BG, fg=WARNING, font=("Segoe UI", 10, "bold"))
    old_warning.place(x=200, y=35)
    start_capslock_monitor(old, old_warning)

    def toggle_old():
        if old.cget("show") == "":
            old.config(show="*")
            btn_old.config(text="Show")
        else:
            old.config(show="")
            btn_old.config(text="Hide")

    btn_old = tk.Button(master, text="Show", command=toggle_old, width=6)
    btn_old.place(x=340, y=8)

    lbl5 = tk.Label(master, text="Enter New Password", bg=CARD_BG, fg=MUTED, font=("Segoe UI", 12, "bold"))
    lbl5.place(x=10, y=60)
    new = tk.Entry(master, width=22, fg="black", relief="solid", font=("Segoe UI", 12, "bold"), show="*")
    new.place(x=200, y=60)
    new_warning = tk.Label(master, text="", bg=CARD_BG, fg=WARNING, font=("Segoe UI", 10, "bold"))
    new_warning.place(x=200, y=85)
    start_capslock_monitor(new, new_warning)

    def toggle_new():
        if new.cget("show") == "":
            new.config(show="*")
            btn_new.config(text="Show")
        else:
            new.config(show="")
            btn_new.config(text="Hide")

    btn_new = tk.Button(master, text="Show", command=toggle_new, width=6)
    btn_new.place(x=340, y=58)

    lbl6 = tk.Label(master, text="Confirm New Password", bg=CARD_BG, fg=MUTED, font=("Segoe UI", 12, "bold"))
    lbl6.place(x=10, y=110)
    nnew = tk.Entry(master, width=22, fg="black", relief="solid", font=("Segoe UI", 12, "bold"), show="*")
    nnew.place(x=200, y=110)
    nnew_warning = tk.Label(master, text="", bg=CARD_BG, fg=WARNING, font=("Segoe UI", 10, "bold"))
    nnew_warning.place(x=200, y=135)
    start_capslock_monitor(nnew, nnew_warning)

    def toggle_nnew():
        if nnew.cget("show") == "":
            nnew.config(show="*")
            btn_nnew.config(text="Show")
        else:
            nnew.config(show="")
            btn_nnew.config(text="Hide")

    btn_nnew = tk.Button(master, text="Show", command=toggle_nnew, width=6)
    btn_nnew.place(x=340, y=108)

    cancel = tk.Button(
        master,
        text="Cancel",
        command=master.destroy,
        fg="white",
        bg=WARNING,
        height=1,
        width=25,
        activebackground="white",
        font=("Segoe UI", 10, "bold"),
    )
    cancel.place(x=200, y=150)

    save1 = tk.Button(
        master,
        text="Save",
        command=save_pass,
        fg="white",
        bg=SUCCESS,
        height=1,
        width=25,
        activebackground="white",
        font=("Segoe UI", 10, "bold"),
    )
    save1.place(x=10, y=150)


def psw():
    assure_path_exists(TRAINING_LABEL_DIR)

    if os.path.isfile(PASSWORD_FILE):
        entered = ask_password_dialog(window, title="Password", prompt="Enter Password", require_confirm=False)
        if entered is None:
            return

        with open(PASSWORD_FILE, "r", encoding="utf-8") as tf:
            key = tf.read()

        if entered == key:
            TrainImages()
        else:
            mess.showerror("Wrong Password", "You have entered wrong password")
    else:
        new_pas = ask_password_dialog(window, title="Set New Password", prompt="Enter new password", require_confirm=True)
        if new_pas is None:
            mess.showerror("No Password Entered", "Password not set!! Please try again")
            return

        with open(PASSWORD_FILE, "w", encoding="utf-8") as tf:
            tf.write(new_pas)

        mess.showinfo("Password Registered", "New password was registered successfully!!")


# ========================= GUI CLEAR =========================
def clear():
    txt.delete(0, "end")
    message1.configure(text="1) Capture Face Images  >>>  2) Train Model")


def clear2():
    txt2.delete(0, "end")
    message1.configure(text="1) Capture Face Images  >>>  2) Train Model")


# ========================= FACE CAPTURE / TRAINING =========================
def TakeImages():
    if not check_haarcascadefile():
        return

    assure_path_exists(TRAINING_IMAGE_DIR)
    ensure_person_details_file()

    try:
        df_existing = pd.read_excel(PERSON_DETAILS_FILE)
        df_existing = df_existing.dropna(how="all")
        if "SERIAL NO." in df_existing.columns:
            serials = pd.to_numeric(df_existing["SERIAL NO."], errors="coerce")
            serial = int(serials.max()) + 1 if serials.notna().any() else 1
        else:
            serial = 1
    except Exception:
        serial = 1

    Id = txt.get().strip()
    name = txt2.get().strip()

    if not Id.isdigit():
        status_label.config(text="ID must be numeric.", fg=WARNING)
        return

    if not name.replace(" ", "").isalpha():
        status_label.config(text="Enter a valid name (letters and spaces only).", fg=WARNING)
        return

    safe_name = name.replace(" ", "_")
    user_folder = os.path.join(TRAINING_IMAGE_DIR, f"{safe_name}_{Id}")
    assure_path_exists(user_folder)

    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        status_label.config(text="Failed to access camera.", fg=WARNING)
        return

    detector = cv2.CascadeClassifier(CASCADE_FILE)
    sampleNum = 0

    try:
        while True:
            ret, img = cam.read()
            if not ret:
                break

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, 1.1, 5)

            for (x, y, w, h) in faces:
                sampleNum += 1
                face_gray = capture_face_roi(gray, x, y, w, h)
                filename = f"{name}.{serial}.{Id}.{sampleNum}.jpg"
                cv2.imwrite(os.path.join(user_folder, filename), face_gray)
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

            cv2.imshow("Capturing Face Images (Press Q to quit)", img)
            key = cv2.waitKey(1) & 0xFF

            if key in [ord("q"), ord("Q")] or sampleNum >= 100:
                break

    finally:
        cam.release()
        cv2.destroyAllWindows()

    try:
        df_row = pd.DataFrame([[serial, int(Id), name]], columns=["SERIAL NO.", "ID", "NAME"])
        try:
            df_all = pd.read_excel(PERSON_DETAILS_FILE)
            df_all = df_all.dropna(how="all")
            df_all = pd.concat([df_all, df_row], ignore_index=True)
        except Exception:
            df_all = df_row

        df_all.to_excel(PERSON_DETAILS_FILE, index=False)
        autofit_excel(PERSON_DETAILS_FILE)
    except Exception as e:
        mess.showerror("Error", f"Failed to save person details: {e}")
        return

    update_registration_counter()
    message1.configure(text="1) Capture Face Images  >>>  2) Train Model (Now click Train Model)")
    status_label.config(text="Face images captured successfully.", fg=SUCCESS)


def getImagesAndLabels(path):
    faces = []
    IDs = []

    for root, _, files in os.walk(path):
        for file in files:
            if not file.lower().endswith((".jpeg", ".jpg", ".png")):
                continue

            imagePath = os.path.join(root, file)

            try:
                pilImage = Image.open(imagePath).convert("L")
                imageNp = np.array(pilImage, "uint8")
                imageNp = cv2.resize(imageNp, (200, 200))
            except Exception:
                continue

            try:
                ID = int(os.path.basename(imagePath).split(".")[2])
            except Exception:
                continue

            faces.append(imageNp)
            IDs.append(ID)

    return faces, IDs


def TrainImages():
    if not check_haarcascadefile():
        return

    assure_path_exists(TRAINING_LABEL_DIR)

    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except AttributeError:
        status_label.config(text="LBPH Recognizer not available. Install opencv-contrib-python.", fg=WARNING)
        return

    faces, IDs = getImagesAndLabels(TRAINING_IMAGE_DIR)

    if len(faces) == 0 or len(IDs) == 0:
        mess.showerror("No Registrations", "Please register someone first!")
        return

    try:
        recognizer.train(faces, np.array(IDs))
        recognizer.save(TRAINER_FILE)
        message1.configure(text="Model trained and saved successfully.")
        update_registration_counter()
        refresh_identification_tree()
        status_label.config(text="Model trained successfully.", fg=SUCCESS)
    except Exception as e:
        status_label.config(text=f"Training failed: {e}", fg=WARNING)


# ========================= IDENTIFICATION =========================
def IdentifyPerson():
    if not check_haarcascadefile():
        return

    assure_path_exists(IDENTIFICATION_DIR)
    ensure_person_details_file()
    ensure_identification_file()

    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except Exception:
        mess.showerror("Error", "LBPH recognizer not available. Please install opencv-contrib-python.")
        return

    if not os.path.isfile(TRAINER_FILE):
        mess.showerror("Data Missing", "Please click on Train Model first.")
        return

    try:
        recognizer.read(TRAINER_FILE)
    except Exception as e:
        mess.showerror("Error", f"Could not read trainer file: {e}")
        return

    if not os.path.isfile(PERSON_DETAILS_FILE):
        mess.showerror("Details Missing", "Person details are missing.")
        return

    try:
        df = pd.read_excel(PERSON_DETAILS_FILE)
        df = df.dropna(how="all")
    except Exception:
        df = pd.DataFrame(columns=["SERIAL NO.", "ID", "NAME"])

    faceCascade = cv2.CascadeClassifier(CASCADE_FILE)
    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        mess.showerror("Error", "Failed to access camera.")
        return

    font = cv2.FONT_HERSHEY_SIMPLEX
    start_time = time.time()
    max_duration = 8
    best_match = None
    logged_ids = set()

    try:
        while True:
            ret, im = cam.read()
            if not ret:
                break

            gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
            faces = faceCascade.detectMultiScale(gray, 1.1, 5)

            for (x, y, w, h) in faces:
                cv2.rectangle(im, (x, y), (x + w, y + h), (34, 197, 94), 2)

                try:
                    face_roi = capture_face_roi(gray, x, y, w, h)
                    predicted_id, conf = recognizer.predict(face_roi)
                except Exception:
                    predicted_id, conf = -1, 100

                label = "Unknown"

                if conf < 70:
                    row = get_student_row_by_id(df, predicted_id)
                    if row is not None:
                        serial_no = row["SERIAL NO."]
                        student_id = row["ID"]
                        student_name = str(row["NAME"])
                        label = student_name

                        candidate = {
                            "serial_no": int(serial_no) if pd.notna(serial_no) else "",
                            "student_id": int(student_id) if pd.notna(student_id) else "",
                            "student_name": student_name,
                            "conf": conf,
                        }

                        if best_match is None or candidate["conf"] < best_match["conf"]:
                            best_match = candidate

                        # Real-time logging: write immediately once per session per person
                        if candidate["student_id"] not in logged_ids:
                            ts = time.time()
                            date_str = datetime.datetime.fromtimestamp(ts).strftime("%d-%m-%Y")
                            time_str = datetime.datetime.fromtimestamp(ts).strftime("%I:%M:%S %p")
                            append_identification_log(candidate["student_id"], candidate["student_name"], date_str, time_str)
                            logged_ids.add(candidate["student_id"])
                            status_label.config(text=f"Identified: {candidate['student_name']}", fg=SUCCESS)
                            window.update_idletasks()

                cv2.putText(im, str(label), (x, y + h + 20), font, 0.9, (255, 255, 255), 2)

            cv2.imshow("Person Identification (Press Q to quit)", im)
            key = cv2.waitKey(1) & 0xFF

            if key in [ord("q"), ord("Q")]:
                break

            if time.time() - start_time >= max_duration:
                break

    finally:
        cam.release()
        cv2.destroyAllWindows()

    if best_match is None:
        status_label.config(text="No valid person was identified.", fg=WARNING)
        return

    refresh_identification_tree()
    status_label.config(text=f"Last detected: {best_match['student_name']}", fg=SUCCESS)


# Backward-compatible alias
TrackImages = IdentifyPerson


# ========================= DELETE FILES =========================
def delete_person_details_xlsx():
    if os.path.exists(PERSON_DETAILS_FILE):
        try:
            os.remove(PERSON_DETAILS_FILE)
            ensure_person_details_file()
            update_registration_counter()
            mess.showinfo("Success", "Person details XLSX file deleted successfully.")
        except Exception as e:
            mess.showerror("Error", f"Failed to delete person details XLSX: {e}")
    else:
        mess.showerror("Error", "Person details XLSX file not found.")


def delete_identification_xlsx():
    if os.path.exists(IDENTIFICATION_FILE):
        try:
            os.remove(IDENTIFICATION_FILE)
            ensure_identification_file()
            refresh_identification_tree()
            mess.showinfo("Success", "Identification log XLSX file deleted successfully.")
        except Exception as e:
            mess.showerror("Error", f"Failed to delete identification log: {e}")
    else:
        mess.showerror("Error", "Identification log XLSX file not found.")


def delete_registered_images():
    if os.path.exists(TRAINING_IMAGE_DIR):
        try:
            shutil.rmtree(TRAINING_IMAGE_DIR)
            assure_path_exists(TRAINING_IMAGE_DIR)
            mess.showinfo("Success", "Registered face images deleted successfully.")
        except Exception as e:
            mess.showerror("Error", f"Failed to delete registered images: {e}")
    else:
        mess.showerror("Error", "TrainingImage folder not found.")


# ========================= CONTACT =========================
def contact():
    mess.showinfo("Contact us", "Please contact us on: dasdarshan7@gmail.com")


# ========================= APP INIT =========================
ensure_person_details_file()
ensure_identification_file()

window = tk.Tk()
window.title("AI Person Identification System")
window.geometry("1280x720")
window.minsize(1180, 680)
window.configure(bg=BG)

style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview", background=TABLE_BG, fieldbackground=TABLE_BG, foreground="#111827", rowheight=28, bordercolor="#cbd5e1", borderwidth=0)
style.configure("Treeview.Heading", background=TABLE_HEADER, foreground="#111827", font=("Segoe UI", 10, "bold"))
style.map("Treeview", background=[("selected", "#bfdbfe")], foreground=[("selected", "#111827")])

# Top header
header = tk.Frame(window, bg=HEADER_BG, height=92)
header.pack(fill="x")
header.pack_propagate(False)

title_label = tk.Label(
    header,
    text="AI Person Identification System",
    bg=HEADER_BG,
    fg=HEADER_FG,
    font=("Segoe UI", 24, "bold"),
)
title_label.pack(pady=(12, 0))

subtitle_label = tk.Label(
    header,
    text="Real-time face recognition and person identification",
    bg=HEADER_BG,
    fg="#dbeafe",
    font=("Segoe UI", 10),
)
subtitle_label.pack()

# Date/time row
info_bar = tk.Frame(window, bg=BG)
info_bar.pack(fill="x", pady=(10, 6))

current_date_label = tk.Label(info_bar, text="", bg="#dbeafe", fg="#0f172a", font=("Segoe UI", 11, "bold"), padx=14, pady=6)
current_date_label.pack(side="left", padx=(20, 10))

clock = tk.Label(info_bar, text="", bg="#dbeafe", fg="#0f172a", font=("Segoe UI", 11, "bold"), padx=14, pady=6)
clock.pack(side="left")


def tick():
    current_time = time.strftime("%I:%M:%S %p")
    clock.config(text=current_time)
    clock.after(1000, tick)


ts = time.time()
date = datetime.datetime.fromtimestamp(ts).strftime("%d-%m-%Y")
day, month, year = date.split("-")
mont = {
    "01": "January",
    "02": "February",
    "03": "March",
    "04": "April",
    "05": "May",
    "06": "June",
    "07": "July",
    "08": "August",
    "09": "September",
    "10": "October",
    "11": "November",
    "12": "December",
}
current_date_label.config(text=f"{day}-{mont[month]}-{year}")
tick()

# Main layout
main = tk.Frame(window, bg=BG)
main.pack(fill="both", expand=True, padx=20, pady=(0, 16))
main.grid_columnconfigure(0, weight=1)
main.grid_columnconfigure(1, weight=1)
main.grid_rowconfigure(0, weight=1)

left_card = tk.Frame(main, bg=CARD_BG, highlightbackground=CARD_BORDER, highlightthickness=1)
left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
left_card.grid_rowconfigure(1, weight=1)

right_card = tk.Frame(main, bg=CARD_BG, highlightbackground=CARD_BORDER, highlightthickness=1)
right_card.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
right_card.grid_rowconfigure(2, weight=1)
right_card.grid_columnconfigure(0, weight=1)

# Left card - registration
left_header = tk.Label(left_card, text="Register New Person", bg="#dbeafe", fg="#0f172a", font=("Segoe UI", 14, "bold"), pady=10)
left_header.pack(fill="x")

reg_body = tk.Frame(left_card, bg=CARD_BG)
reg_body.pack(fill="both", expand=True, padx=18, pady=16)

id_label = tk.Label(reg_body, text="Enter ID", bg=CARD_BG, fg="#111827", font=("Segoe UI", 12, "bold"))
id_label.grid(row=0, column=0, sticky="w")
txt = tk.Entry(reg_body, width=28, font=("Segoe UI", 12))
txt.grid(row=1, column=0, sticky="ew", pady=(6, 14))
txt_warning = tk.Label(reg_body, text="", fg=WARNING, bg=CARD_BG, font=("Segoe UI", 9, "bold"))
txt_warning.grid(row=2, column=0, sticky="w")
start_capslock_monitor(txt, txt_warning)

name_label = tk.Label(reg_body, text="Enter Name", bg=CARD_BG, fg="#111827", font=("Segoe UI", 12, "bold"))
name_label.grid(row=3, column=0, sticky="w", pady=(12, 0))
txt2 = tk.Entry(reg_body, width=28, font=("Segoe UI", 12))
txt2.grid(row=4, column=0, sticky="ew", pady=(6, 14))
txt2_warning = tk.Label(reg_body, text="", fg=WARNING, bg=CARD_BG, font=("Segoe UI", 9, "bold"))
txt2_warning.grid(row=5, column=0, sticky="w")
start_capslock_monitor(txt2, txt2_warning)

message1 = tk.Label(
    reg_body,
    text="1) Capture Face Images  >>>  2) Train Model",
    bg="#eff6ff",
    fg="#0f172a",
    font=("Segoe UI", 10, "bold"),
    pady=10,
)
message1.grid(row=6, column=0, sticky="ew", pady=(16, 10))

status_label = tk.Label(reg_body, text="", bg=CARD_BG, fg=MUTED, font=("Segoe UI", 10, "bold"), wraplength=420, justify="left")
status_label.grid(row=7, column=0, sticky="w", pady=(0, 10))

btn_row = tk.Frame(reg_body, bg=CARD_BG)
btn_row.grid(row=8, column=0, sticky="ew", pady=(10, 0))
btn_row.grid_columnconfigure((0, 1), weight=1)

clear_btn = tk.Button(btn_row, text="Clear ID", command=clear, bg="#e2e8f0", fg="#0f172a", font=("Segoe UI", 10, "bold"), relief="flat", padx=12, pady=8)
clear_btn.grid(row=0, column=0, sticky="ew", padx=(0, 6))

clear_btn2 = tk.Button(btn_row, text="Clear Name", command=clear2, bg="#e2e8f0", fg="#0f172a", font=("Segoe UI", 10, "bold"), relief="flat", padx=12, pady=8)
clear_btn2.grid(row=0, column=1, sticky="ew", padx=(6, 0))

capture_button = tk.Button(reg_body, text="Capture Face Images", command=TakeImages, bg=ACCENT, fg="white", font=("Segoe UI", 11, "bold"), relief="flat", padx=16, pady=10)
capture_button.grid(row=9, column=0, sticky="ew", pady=(18, 10))

train_button = tk.Button(reg_body, text="Train Model", command=psw, bg=ACCENT_DARK, fg="white", font=("Segoe UI", 11, "bold"), relief="flat", padx=16, pady=10)
train_button.grid(row=10, column=0, sticky="ew", pady=(0, 14))

stats_label = tk.Label(reg_body, text="", bg=CARD_BG, fg="#111827", font=("Segoe UI", 11, "bold"))
stats_label.grid(row=11, column=0, sticky="w", pady=(6, 0))
update_registration_counter()

# Right card - live identification and logs
right_header = tk.Label(right_card, text="Live Identification", bg="#dbeafe", fg="#0f172a", font=("Segoe UI", 14, "bold"), pady=10)
right_header.pack(fill="x")

controls = tk.Frame(right_card, bg=CARD_BG)
controls.pack(fill="x", padx=18, pady=(16, 8))

identify_button = tk.Button(controls, text="Identify Person", command=IdentifyPerson, bg=SUCCESS, fg="white", font=("Segoe UI", 11, "bold"), relief="flat", padx=16, pady=10)
identify_button.pack(side="left")

quitWindow = tk.Button(controls, text="Quit", command=window.destroy, bg=WARNING, fg="white", font=("Segoe UI", 11, "bold"), relief="flat", padx=16, pady=10)
quitWindow.pack(side="right")

log_title = tk.Label(right_card, text="Identification Log", bg=CARD_BG, fg="#111827", font=("Segoe UI", 12, "bold"))
log_title.pack(anchor="w", padx=18, pady=(8, 6))

log_frame = tk.Frame(right_card, bg=CARD_BG)
log_frame.pack(fill="both", expand=True, padx=18, pady=(0, 12))
log_frame.grid_rowconfigure(0, weight=1)
log_frame.grid_columnconfigure(0, weight=1)

tv = ttk.Treeview(log_frame, height=16, columns=("id", "name", "date", "time"))
tv.grid(row=0, column=0, sticky="nsew")

tv.column("#0", width=70, anchor=tk.CENTER)
tv.column("id", width=80, anchor=tk.CENTER)
tv.column("name", width=160, anchor=tk.CENTER)
tv.column("date", width=110, anchor=tk.CENTER)
tv.column("time", width=110, anchor=tk.CENTER)

tv.heading("#0", text="Serial No.")
tv.heading("id", text="ID")
tv.heading("name", text="Name")
tv.heading("date", text="Date")
tv.heading("time", text="Time")

scroll = ttk.Scrollbar(log_frame, orient="vertical", command=tv.yview)
scroll.grid(row=0, column=1, sticky="ns")
tv.configure(yscrollcommand=scroll.set)

scroll_x = ttk.Scrollbar(log_frame, orient="horizontal", command=tv.xview)
scroll_x.grid(row=1, column=0, sticky="ew")
tv.configure(xscrollcommand=scroll_x.set)

# Footer controls
footer = tk.Frame(window, bg=BG)
footer.pack(fill="x", padx=20, pady=(0, 12))

footer_left = tk.Frame(footer, bg=BG)
footer_left.pack(side="left")

help_btn = tk.Menubutton(footer_left, text="Help", bg="#dbeafe", fg="#0f172a", relief="flat", font=("Segoe UI", 10, "bold"), padx=12, pady=8)
help_menu = tk.Menu(help_btn, tearoff=0)
help_menu.add_command(label="Change Password", command=change_pass)
help_menu.add_command(label="Contact Us", command=contact)
help_btn.config(menu=help_menu)
help_btn.pack(side="left", padx=(0, 8))

reg_delete_btn = tk.Button(footer_left, text="Delete Person XLSX", command=delete_person_details_xlsx, bg="#ef4444", fg="white", font=("Segoe UI", 10, "bold"), relief="flat", padx=12, pady=8)
reg_delete_btn.pack(side="left", padx=(0, 8))

log_delete_btn = tk.Button(footer_left, text="Delete Log XLSX", command=delete_identification_xlsx, bg="#ef4444", fg="white", font=("Segoe UI", 10, "bold"), relief="flat", padx=12, pady=8)
log_delete_btn.pack(side="left", padx=(0, 8))

img_delete_btn = tk.Button(footer_left, text="Delete Registered Images", command=delete_registered_images, bg="#ef4444", fg="white", font=("Segoe UI", 10, "bold"), relief="flat", padx=12, pady=8)
img_delete_btn.pack(side="left")

# Remove old Gmail widgets by not creating them at all.

refresh_identification_tree()
window.mainloop()
